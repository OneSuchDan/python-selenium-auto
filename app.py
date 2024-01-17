from files_otros.store_selenium import StoreSele
import openpyxl
import logging

# Se lee el archivo excel con datos de ejemplo
excel_dataframe = openpyxl.load_workbook("ejemplo.xlsx")
dataframe = excel_dataframe.active
data = []
# Se itera en filas y columnas de 1 a 5 columnas nombre completo,seccion,codigo credencial, numero y mail (opcional)
# - Ignora la primera fila que va contener los titulos
# - La información se almacenara en la variable {data}
for row in range(1, dataframe.max_row):
    _row = [row, ]
    for col in dataframe.iter_cols(1, 5):
        _row.append(col[row].value)

    data.append(_row)
# A partir de la data obtenida enterior se utiliza la clase para hacer la automatización
#   - log de archivos subidos con formato de fecha
#   - Se itera en la informacion del archivo de excel y se remueve el primer elemento del archivo de excel
with StoreSele() as bot:
    logging.basicConfig(filename='subidos.log',format='%(asctime)s %(message)s',datefmt='%m/%d/%Y %I:%M:%S %p', encoding='utf-8', level=logging.INFO)
    logging.info("Iniciado")
    for arreglo in enumerate(data):
        arreglo.pop(0)
        name,section,code,phone,mail=arreglo
        if name is None or section is None or code is None or phone is None:
            logging.error(f'{name} No fue insertado por datos faltantes')
            continue
        bot.get_url()
        bot.get_error(name)
        bot.rellenar_form(name,section,code,phone,mail)
